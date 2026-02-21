"""
Charity Prospector — Web App
Finds qualifying charities via ProPublica Nonprofit Explorer API,
checks Form 990 data for fundraising expenses and Schedule G agencies,
and builds a downloadable Excel contact list.

Deploy on Replit: just click Run!
"""

import streamlit as st
import requests
import json
import xml.etree.ElementTree as ET
import time
import io
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl not installed. Add it to requirements.txt")
    st.stop()

# ─── Configuration ───────────────────────────────────────────────────────────
BASE_URL = "https://projects.propublica.org/nonprofits/api/v2"
SEARCH_URL = f"{BASE_URL}/search.json"
ORG_URL = f"{BASE_URL}/organizations"
XML_URL = "https://projects.propublica.org/nonprofits/download-xml"
REQUEST_DELAY = 0.5

# Apollo.io API
APOLLO_API_URL = "https://api.apollo.io/api/v1"

BROAD_SEARCH_KEYWORDS = [
    "foundation", "hospital", "university", "association", "society",
    "institute", "museum", "community", "health", "education",
    "children", "medical", "research", "services", "arts",
    "wildlife", "conservation", "relief", "humanitarian", "scholarship",
    "veterans", "housing", "faith", "church", "mission",
    "development", "advocacy", "prevention", "counseling", "food bank",
]


# ─── API Helpers ─────────────────────────────────────────────────────────────
@st.cache_data(ttl=3600, show_spinner=False)
def api_get(url, params_tuple=None):
    params = dict(params_tuple) if params_tuple else None
    for attempt in range(3):
        try:
            time.sleep(REQUEST_DELAY)
            resp = requests.get(url, params=params, timeout=30,
                                headers={"User-Agent": "CharityProspector/1.0"})
            if resp.status_code == 200:
                return resp.json()
            elif resp.status_code == 429:
                time.sleep(10 * (attempt + 1))
            else:
                return None
        except Exception:
            time.sleep(5)
    return None


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_xml(url):
    for attempt in range(4):
        try:
            time.sleep(max(REQUEST_DELAY, 1.5))
            resp = requests.get(url, timeout=30,
                                headers={"User-Agent": "CharityProspector/1.0"})
            if resp.status_code == 200:
                content = resp.content
                if content.startswith(b'\xef\xbb\xbf'):
                    content = content[3:]
                if content.lstrip().startswith(b'Error 429') or content.lstrip().startswith(b'<html'):
                    time.sleep(15 * (attempt + 1))
                    continue
                return content
            elif resp.status_code == 429:
                time.sleep(15 * (attempt + 1))
            else:
                return None
        except Exception:
            time.sleep(5)
    return None


def search_orgs(query="", state=None, page=0):
    params = {"q": query, "page": str(page)}
    if state:
        params["state"] = state
    return api_get(SEARCH_URL, params_tuple=tuple(sorted(params.items())))


def get_org_details(ein):
    url = f"{ORG_URL}/{ein}.json"
    return api_get(url)


def get_xml_url(org_data):
    org_info = org_data.get("organization", {})
    object_id = org_info.get("latest_object_id")
    if object_id:
        return f"{XML_URL}?object_id={object_id}"
    return None


# ─── Revenue Check (from JSON API) ──────────────────────────────────────────
def check_revenue(org_data, min_rev, max_rev):
    if not org_data or "filings_with_data" not in org_data:
        return False, 0, 0
    filings = org_data.get("filings_with_data", [])
    if not filings:
        return False, 0, 0
    filing = filings[0]
    revenue = filing.get("totrevenue") or filing.get("totrevnue") or filing.get("totrcptperbks") or 0
    total_expenses = filing.get("totfuncexpns", 0) or 0
    revenue = revenue or 0
    if min_rev <= revenue <= max_rev:
        return True, revenue, total_expenses
    return False, revenue, total_expenses


# ─── Fundraising Expense from XML ───────────────────────────────────────────
def get_fundraising_expense_from_xml(xml_content):
    if not xml_content:
        return 0
    try:
        root = ET.fromstring(xml_content)
        for elem in root.iter():
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            if tag in ('CYTotalFundraisingExpenseAmt', 'TotalFundrsngExpCurrentYrAmt'):
                if elem.text:
                    try:
                        return float(elem.text)
                    except ValueError:
                        pass
        for elem in root.iter():
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            if tag == 'FundraisingAmt':
                parent = None
                for p in root.iter():
                    if elem in list(p):
                        ptag = p.tag.split('}')[-1] if '}' in p.tag else p.tag
                        if 'Total' in ptag:
                            if elem.text:
                                try:
                                    return float(elem.text)
                                except ValueError:
                                    pass
    except Exception:
        pass
    return 0


def build_org_details(org_data, revenue, total_expenses, fundraising_exp, xml_url):
    org_info = org_data.get("organization", {})
    filings = org_data.get("filings_with_data", [])
    filing = filings[0] if filings else {}
    return {
        "ein": str(org_info.get("ein", "")),
        "name": org_info.get("name", ""),
        "city": org_info.get("city", ""),
        "state": org_info.get("state", ""),
        "ntee_code": org_info.get("ntee_code", ""),
        "subsection": org_info.get("subseccd", ""),
        "revenue": revenue,
        "total_expenses": total_expenses,
        "fundraising_expenses": fundraising_exp,
        "tax_year": filing.get("tax_prd_yr") or filing.get("tax_prd", ""),
        "fiscal_year_end": filing.get("prd_end") or filing.get("tax_prd", ""),
        "form_type": filing.get("formtype", ""),
        "filing_url": filing.get("pdf_url", ""),
        "xml_url": xml_url or "",
        "updated": filing.get("updated", ""),
    }


# ─── Schedule G Parser ───────────────────────────────────────────────────────
def parse_schedule_g_from_content(xml_content):
    if not xml_content:
        return []
    try:
        root = ET.fromstring(xml_content)
        agencies = []

        # Find fundraiser activity entries (multiple schema versions)
        found_entries = []
        for elem in root.iter():
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            if tag in ('FundraiserActivityInfoGrp', 'ProfessionalFundraising',
                       'FundraisingActivityGroup', 'ProfFundRaisingGrp'):
                found_entries.append(elem)

        for entry in found_entries:
            agency = {}

            # Agency name
            for name_tag in ('PersonNm', 'BusinessNameLine1Txt', 'BusinessNameLine1',
                             'BusinessName', 'OrganizationBusinessName'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == name_tag and child.text:
                        agency['name'] = child.text.strip()
                        break
                if 'name' in agency:
                    break

            if 'name' not in agency:
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == 'BusinessNameLine1Txt' and child.text:
                        agency['name'] = child.text.strip()
                        break

            # Amount paid
            for amt_tag in ('RetainedByContractorAmt', 'AmtPaidToFundraiser',
                            'CompensationAmount', 'AmountPaidToFundraiser', 'CompensationAmt'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == amt_tag and child.text:
                        try:
                            agency['amount_paid'] = float(child.text)
                        except ValueError:
                            pass
                        break
                if 'amount_paid' in agency:
                    break

            # Amount raised
            for raised_tag in ('GrossReceiptsFromActivityAmt', 'AmountRaisedByContractor',
                               'GrossReceiptsFromActivity'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == raised_tag and child.text:
                        try:
                            agency['amount_raised'] = float(child.text)
                        except ValueError:
                            pass
                        break
                if 'amount_raised' in agency:
                    break

            # Activity
            for desc_tag in ('ActivityTxt', 'Activity', 'Description'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == desc_tag and child.text:
                        agency['activity'] = child.text.strip()
                        break
                if 'activity' in agency:
                    break

            # Address
            for addr_tag in ('CityNm', 'City'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == addr_tag and child.text:
                        agency['city'] = child.text.strip()
                        break
                if 'city' in agency:
                    break

            for state_tag in ('StateAbbreviationCd', 'State'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == state_tag and child.text:
                        agency['state'] = child.text.strip()
                        break
                if 'state' in agency:
                    break

            if agency.get('name'):
                agencies.append(agency)

        return agencies
    except Exception:
        return []


# ─── Officer/Contact Extraction ─────────────────────────────────────────────
def extract_officers_from_xml(xml_content):
    if not xml_content:
        return []
    try:
        root = ET.fromstring(xml_content)
        officers = []

        found_entries = []
        for elem in root.iter():
            tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            if tag in ('Form990PartVIISectionAGrp', 'OfficerDirectorTrusteeEmplGrp',
                       'CompensationInfoGrp', 'Form990PartVIISectionA'):
                found_entries.append(elem)

        for entry in found_entries:
            person = {}
            for name_tag in ('PersonNm', 'PersonFullName', 'Name',
                             'BusinessNameLine1Txt', 'BusinessNameLine1'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == name_tag and child.text:
                        person['name'] = child.text.strip()
                        break
                if 'name' in person:
                    break

            for title_tag in ('TitleTxt', 'Title', 'PersonTitleTxt'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == title_tag and child.text:
                        person['title'] = child.text.strip()
                        break
                if 'title' in person:
                    break

            for comp_tag in ('ReportableCompFromOrgAmt', 'ReportableCompFromOrg',
                             'CompensationAmount', 'TotalCompensation'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == comp_tag and child.text:
                        try:
                            person['compensation'] = float(child.text)
                        except ValueError:
                            pass
                        break
                if 'compensation' in person:
                    break

            for hours_tag in ('AverageHoursPerWeekRt', 'AverageHoursPerWeek',
                              'AvgHoursPerWkDevotedToPosRt'):
                for child in entry.iter():
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if ctag == hours_tag and child.text:
                        person['hours_per_week'] = child.text.strip()
                        break
                if 'hours_per_week' in person:
                    break

            if person.get('name'):
                officers.append(person)
        return officers
    except Exception:
        return []


def filter_fundraising_contacts(officers):
    """Score and rank officers for fundraising/development relevance. Return top 4."""
    fundraising_kw = ['development', 'fundrais', 'advancement', 'donor', 'philanthrop',
                      'annual giving', 'major gift', 'planned giving', 'campaign',
                      'chief development', 'cdo', 'vp develop', 'vice president develop']
    leadership_kw = ['chief', 'president', 'executive director', 'ceo', 'cfo', 'coo',
                     'vp', 'vice president', 'svp', 'evp', 'director', 'secretary', 'treasurer']

    scored = []
    for officer in officers:
        title = (officer.get('title', '') or '').lower()
        score = 0
        for kw in fundraising_kw:
            if kw in title:
                score += 10
                break
        for kw in leadership_kw:
            if kw in title:
                score += 5
                break
        if officer.get('compensation', 0) and officer['compensation'] > 0:
            score += 3
        try:
            if float(officer.get('hours_per_week', '0')) >= 30:
                score += 2
        except:
            pass
        if score > 0:
            officer['relevance_score'] = score
            officer['source'] = 'Form 990'
            scored.append(officer)

    scored.sort(key=lambda x: (x.get('relevance_score', 0), x.get('compensation', 0)), reverse=True)
    return scored[:4]


# ─── Apollo.io Contact Enrichment ────────────────────────────────────────────

def apollo_search_contacts(org_name, apollo_key):
    """
    Search Apollo.io for people at an organization matching fundraising/development titles.
    Returns list of contacts with name, title, email, LinkedIn, phone.
    """
    if not apollo_key:
        return []

    titles = [
        "VP Development", "Vice President Development",
        "Chief Development Officer", "CDO",
        "Director of Development", "Director of Fundraising",
        "Senior Director Development", "SVP Development",
        "Executive Director", "CEO", "President",
    ]

    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
    }

    contacts = []

    try:
        org_search_url = f"{APOLLO_API_URL}/mixed_people/search"
        payload = {
            "api_key": apollo_key,
            "q_organization_name": org_name,
            "person_titles": titles,
            "page": 1,
            "per_page": 5,
        }

        time.sleep(0.3)
        resp = requests.post(org_search_url, json=payload, headers=headers, timeout=30)

        if resp.status_code == 200:
            data = resp.json()
            people = data.get("people", [])

            for person in people[:4]:
                contact = {
                    "name": f"{person.get('first_name', '')} {person.get('last_name', '')}".strip(),
                    "title": person.get("title", ""),
                    "email": person.get("email", ""),
                    "linkedin_url": person.get("linkedin_url", ""),
                    "phone": "",
                    "compensation": 0,
                    "hours_per_week": "",
                    "relevance_score": 8,
                    "source": "Apollo.io",
                }

                phone_numbers = person.get("phone_numbers", [])
                if phone_numbers:
                    contact["phone"] = phone_numbers[0].get("sanitized_number", "")

                if contact["name"]:
                    contacts.append(contact)

    except Exception:
        pass

    return contacts


# ─── Excel Builder ───────────────────────────────────────────────────────────
def build_excel(charities, all_contacts, params):
    """Build formatted Excel workbook and return as bytes."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    currency_fmt = '$#,##0'

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Charity Summary"
    headers = ["EIN", "Organization Name", "City", "State", "NTEE Code",
               "Total Revenue", "Total Expenses", "Fundraising Expenses",
               "Tax Year", "Fiscal Year End", "# Agencies", "Top Agency",
               "Top Agency Spend", "# Contacts"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border

    for ri, ch in enumerate(charities, 2):
        agencies = ch.get("agencies", [])
        contacts = all_contacts.get(ch["ein"], [])
        top = agencies[0] if agencies else {}
        vals = [ch["ein"], ch["name"], ch["city"], ch["state"], ch["ntee_code"],
                ch["revenue"], ch["total_expenses"], ch["fundraising_expenses"],
                ch["tax_year"], ch["fiscal_year_end"], len(agencies),
                top.get("name", "N/A"), top.get("amount_paid", 0), len(contacts)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=col, value=v)
            c.border = thin_border
            if col in (6, 7, 8, 13):
                c.number_format = currency_fmt

    for col in range(1, len(headers) + 1):
        mx = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(charities) + 2))
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 4, 35)

    # ── Sheet 2: Agencies ──
    ws2 = wb.create_sheet("Fundraising Agencies")
    ah = ["EIN", "Organization", "Agency Name", "Agency City", "Agency State",
          "Amount Paid", "Amount Raised", "Activity"]
    for col, h in enumerate(ah, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border
    ri = 2
    for ch in charities:
        for ag in ch.get("agencies", []):
            vals = [ch["ein"], ch["name"], ag.get("name", ""), ag.get("city", ""),
                    ag.get("state", ""), ag.get("amount_paid", 0),
                    ag.get("amount_raised", 0), ag.get("activity", "")]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=ri, column=col, value=v)
                c.border = thin_border
                if col in (6, 7):
                    c.number_format = currency_fmt
            ri += 1
    for col in range(1, len(ah) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 25

    # ── Sheet 3: Contacts ──
    ws3 = wb.create_sheet("Contacts")
    ch_h = ["EIN", "Organization", "Contact Name", "Title", "Compensation",
            "Hours/Week", "Relevance Score", "Email", "LinkedIn", "Phone", "Source"]
    for col, h in enumerate(ch_h, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border
    ri = 2
    for ch in charities:
        for ct in all_contacts.get(ch["ein"], []):
            vals = [ch["ein"], ch["name"], ct.get("name", ""), ct.get("title", ""),
                    ct.get("compensation", 0), ct.get("hours_per_week", ""),
                    ct.get("relevance_score", 0), ct.get("email", ""),
                    ct.get("linkedin_url", ""), ct.get("phone", ""),
                    ct.get("source", "Form 990")]
            for col, v in enumerate(vals, 1):
                c = ws3.cell(row=ri, column=col, value=v)
                c.border = thin_border
                if col == 5:
                    c.number_format = currency_fmt
            ri += 1
    for col in range(1, len(ch_h) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # ── Sheet 4: Criteria ──
    ws4 = wb.create_sheet("Criteria & Notes")
    notes = [
        ("Parameter", "Value"),
        ("Revenue Range", f"${params['min_rev']/1e6:.0f}M - ${params['max_rev']/1e6:.0f}M"),
        ("Min Fundraising Expense", f"${params['min_fund']/1e6:.0f}M"),
        ("Min Agency Spend (Schedule G)", f"${params['min_agency']/1e3:.0f}K"),
        ("Organization Type", "501(c)(3)"),
        ("Target Contacts", "3-4 Fundraising/Development leaders per org"),
        ("", ""),
        ("Data Sources", ""),
        ("Financial Data", "ProPublica Nonprofit Explorer API v2"),
        ("Schedule G / Agencies", "IRS Form 990 XML E-Files"),
        ("Officer/Contact Data", "Form 990 Part VII Section A"),
        ("Contact Enrichment", "Apollo.io API (if key provided)"),
        ("", ""),
        ("Notes", ""),
        ("", "Contacts sourced from both Form 990 and Apollo.io where available"),
        ("", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
    ]
    for ri, (k, v) in enumerate(notes, 1):
        ws4.cell(row=ri, column=1, value=k).font = Font(bold=True) if k else Font()
        ws4.cell(row=ri, column=2, value=v)
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# STREAMLIT APP
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Charity Prospector", page_icon="🔍", layout="wide")

st.title("🔍 Charity Prospector")
st.markdown("Find qualifying charities using ProPublica Nonprofit Explorer data, "
            "verify fundraising expenses and Schedule G agencies, and build a contact list.")

# ─── Sidebar: Parameters ─────────────────────────────────────────────────────
with st.sidebar:
    st.header("Search Parameters")

    min_rev = st.number_input("Min Revenue ($)", value=20_000_000, step=1_000_000, format="%d")
    max_rev = st.number_input("Max Revenue ($)", value=200_000_000, step=1_000_000, format="%d")
    min_fund = st.number_input("Min Fundraising Expense ($)", value=2_000_000, step=100_000, format="%d")
    min_agency = st.number_input("Min Agency Spend on Schedule G ($)", value=500_000, step=50_000, format="%d")
    target_count = st.number_input("Target # of qualifying charities", value=10, min_value=1, max_value=100)

    st.divider()
    st.header("Optional Filters")
    state_filter = st.text_input("State (2-letter code, leave blank for all)", value="", max_chars=2).upper() or None
    search_query = st.text_input("Keyword search (leave blank to auto-cycle common terms)", value="",
                                  help="If blank, the app cycles through keywords like 'foundation', 'hospital', 'university', etc. to find diverse charities.")

    st.divider()
    max_pages = st.slider("Max search pages to scan", 10, 500, 200,
                          help="Each page has 25 orgs. More pages = more thorough but slower.")

    st.divider()
    st.header("👤 Contact Enrichment")
    apollo_key = st.text_input("Apollo.io API Key (optional)",
                               type="password",
                               help="Free at apollo.io — 50 credits/month. "
                                    "Adds email, LinkedIn URL, and phone for contacts. "
                                    "Leave blank to skip enrichment.")
    use_apollo = bool(apollo_key)
    if use_apollo:
        st.success("Apollo.io enabled — contacts will include email, LinkedIn, phone")

# ─── Main area ───────────────────────────────────────────────────────────────
params = {"min_rev": min_rev, "max_rev": max_rev, "min_fund": min_fund, "min_agency": min_agency}

# Session state for results
if "qualifying" not in st.session_state:
    st.session_state.qualifying = []
if "all_contacts" not in st.session_state:
    st.session_state.all_contacts = {}
if "running" not in st.session_state:
    st.session_state.running = False
if "phase" not in st.session_state:
    st.session_state.phase = "idle"

col1, col2, col3 = st.columns(3)
with col1:
    start_search = st.button("🚀 Start Search", type="primary", use_container_width=True)
with col2:
    get_contacts = st.button("👤 Get Contacts", use_container_width=True,
                             disabled=len(st.session_state.qualifying) == 0)
with col3:
    download_ready = len(st.session_state.qualifying) > 0

# ─── Phase 1: Search ────────────────────────────────────────────────────────
if start_search:
    st.session_state.qualifying = []
    st.session_state.all_contacts = {}

    qualifying = []
    checked = 0
    revenue_match = 0

    progress_bar = st.progress(0, text="Starting search...")
    status = st.empty()
    results_area = st.empty()
    log_area = st.expander("Detailed log", expanded=False)
    log_lines = []

    def add_log(msg):
        log_lines.append(f"`{datetime.now().strftime('%H:%M:%S')}` {msg}")
        with log_area:
            st.markdown("\n\n".join(log_lines[-50:]))

    seen_eins = set()
    api_errors = 0

    if search_query:
        search_keywords = [search_query]
    else:
        search_keywords = list(BROAD_SEARCH_KEYWORDS)

    pages_per_keyword = max(1, max_pages // len(search_keywords))

    for kw_idx, keyword in enumerate(search_keywords):
        if len(qualifying) >= target_count:
            break

        for page in range(pages_per_keyword):
            if len(qualifying) >= target_count:
                break

            overall_page = kw_idx * pages_per_keyword + page
            pct = min(len(qualifying) / target_count, 0.99)
            progress_bar.progress(pct, text=f"Keyword: \"{keyword}\" p{page} | Checked: {checked} | Qualified: {len(qualifying)}/{target_count}")

            data = search_orgs(query=keyword, state=state_filter, page=page)
            if not data or "organizations" not in data:
                api_errors += 1
                add_log(f"Keyword \"{keyword}\" page {page}: API error or no results.")
                if api_errors >= 5:
                    add_log(f"⚠️ Too many API errors. The ProPublica search API may be experiencing issues.")
                break

            orgs = data.get("organizations", [])
            if not orgs:
                add_log(f"Keyword \"{keyword}\" page {page}: No more results.")
                break

            for org in orgs:
                if len(qualifying) >= target_count:
                    break

                ein = org.get("ein")
                if ein in seen_eins:
                    continue
                seen_eins.add(ein)

                name = org.get("name", "Unknown")
                checked += 1
                status.info(f"[{checked}] Checking: {name}...")

                org_data = get_org_details(ein)
                if not org_data:
                    continue

                in_range, revenue, total_expenses = check_revenue(org_data, min_rev, max_rev)
                if not in_range:
                    continue

                revenue_match += 1
                add_log(f"Checking: **{name}** (EIN: {ein}) — Rev: ${revenue:,.0f}")

                xml_url = get_xml_url(org_data)
                if not xml_url:
                    add_log(f"  ❌ No XML e-file available")
                    continue

                xml_content = fetch_xml(xml_url)
                if not xml_content:
                    add_log(f"  ❌ Could not download XML e-file")
                    continue

                fundraising_exp = get_fundraising_expense_from_xml(xml_content)
                if fundraising_exp < min_fund:
                    add_log(f"  ❌ Fundraising expense ${fundraising_exp:,.0f} < ${min_fund:,.0f}")
                    continue

                add_log(f"  ✅ Revenue: ${revenue:,.0f} | Fundraising: ${fundraising_exp:,.0f}")
                add_log(f"  🔍 Parsing Schedule G for agencies...")

                agencies = parse_schedule_g_from_content(xml_content)

                if not agencies:
                    add_log(f"  ❌ No Schedule G agency data found")
                    continue

                qualifying_agencies = [a for a in agencies if a.get("amount_paid", 0) >= min_agency]

                if not qualifying_agencies:
                    agency_summary = ", ".join(f"{a.get('name','?')} (${a.get('amount_paid',0):,.0f})" for a in agencies[:3])
                    add_log(f"  ❌ Agencies found but none >= ${min_agency:,.0f}: {agency_summary}")
                    continue

                details = build_org_details(org_data, revenue, total_expenses, fundraising_exp, xml_url)
                details["agencies"] = qualifying_agencies

                for a in qualifying_agencies:
                    add_log(f"  💰 Agency: **{a.get('name', 'Unknown')}** — Paid: ${a.get('amount_paid', 0):,.0f}")

                qualifying.append(details)
                add_log(f"  🎯 **QUALIFIED #{len(qualifying)}: {name}**")

                with results_area.container():
                    st.subheader(f"Qualifying Charities ({len(qualifying)}/{target_count})")
                    display_data = []
                    for q in qualifying:
                        top_ag = q["agencies"][0] if q["agencies"] else {}
                        display_data.append({
                            "Name": q["name"],
                            "State": q["state"],
                            "Revenue": f"${q['revenue']:,.0f}",
                            "Fundraising $": f"${q['fundraising_expenses']:,.0f}",
                            "Top Agency": top_ag.get("name", "N/A"),
                            "Agency Spend": f"${top_ag.get('amount_paid', 0):,.0f}",
                            "Tax Year": q["tax_year"],
                        })
                    st.dataframe(display_data, use_container_width=True)

    progress_bar.progress(1.0, text=f"Done! Found {len(qualifying)} qualifying charities from {checked} checked.")
    status.success(f"Search complete. {len(qualifying)} charities qualified out of {checked} checked ({revenue_match} in revenue range).")
    st.session_state.qualifying = qualifying

# ─── Phase 2: Contacts ──────────────────────────────────────────────────────
if get_contacts and st.session_state.qualifying:
    charities = st.session_state.qualifying
    all_contacts = {}

    progress = st.progress(0, text="Extracting contacts...")
    contact_status = st.empty()

    for i, ch in enumerate(charities):
        progress.progress((i + 1) / len(charities), text=f"Getting contacts for {ch['name']}...")
        contact_status.info(f"[{i+1}/{len(charities)}] Parsing Form 990 for {ch['name']}...")

        contacts = []

        # Source 1: Form 990 Part VII officers
        xml_url = ch.get("xml_url", "")
        xml_content = fetch_xml(xml_url) if xml_url else None
        officers = extract_officers_from_xml(xml_content)
        filtered_990 = filter_fundraising_contacts(officers)
        contacts.extend(filtered_990)

        # Source 2: Apollo.io enrichment (if key provided)
        if use_apollo:
            contact_status.info(f"[{i+1}/{len(charities)}] Apollo.io lookup for {ch['name']}...")
            apollo_contacts = apollo_search_contacts(ch["name"], apollo_key)
            for ac in apollo_contacts:
                # Avoid duplicates by name
                existing_names = {c.get("name", "").lower() for c in contacts}
                if ac.get("name", "").lower() not in existing_names:
                    contacts.append(ac)

        # Keep top 4
        contacts = contacts[:4]
        all_contacts[ch["ein"]] = contacts

    st.session_state.all_contacts = all_contacts
    total = sum(len(c) for c in all_contacts.values())
    progress.progress(1.0, text=f"Done! Found {total} contacts across {len(charities)} charities.")
    contact_status.success(f"Extracted {total} contacts.")

    # Show contacts
    st.subheader("Contacts Found")
    for ch in charities:
        contacts = all_contacts.get(ch["ein"], [])
        if contacts:
            st.markdown(f"**{ch['name']}** ({ch['state']})")
            for ct in contacts:
                src = ct.get("source", "Form 990")
                email = ct.get("email", "")
                li = ct.get("linkedin_url", "")
                extras = []
                if email:
                    extras.append(f"📧 {email}")
                if li:
                    extras.append(f"🔗 [LinkedIn]({li})")
                extras_str = " | ".join(extras) if extras else ""
                st.markdown(f"- {ct.get('name', 'N/A')} — *{ct.get('title', 'N/A')}* "
                            f"[{src}] {extras_str}")

# ─── Download Button ─────────────────────────────────────────────────────────
if st.session_state.qualifying:
    st.divider()
    st.subheader("📥 Download Results")

    excel_buf = build_excel(st.session_state.qualifying, st.session_state.all_contacts, params)
    st.download_button(
        label="Download Excel Spreadsheet",
        data=excel_buf,
        file_name=f"charity_prospector_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    # Also show JSON download
    json_str = json.dumps(st.session_state.qualifying, indent=2)
    st.download_button(
        label="Download Raw JSON Data",
        data=json_str,
        file_name="qualifying_charities.json",
        mime="application/json",
    )

# ─── Info Footer ─────────────────────────────────────────────────────────────
st.divider()
with st.expander("ℹ️ How this works"):
    st.markdown("""
**Data Flow:**
1. Searches ProPublica's Nonprofit Explorer API for organizations
2. Checks revenue from detailed filing data (organization endpoint)
3. Downloads IRS Form 990 XML e-files to extract fundraising expenses
4. Parses **Schedule G** from the XML for professional fundraising agencies and amounts
5. Filters for agencies receiving ≥ the minimum spend threshold
6. Extracts officer/key employee data from Form 990 **Part VII**
7. Optionally enriches contacts via **Apollo.io** (email, LinkedIn, phone)
8. Packages everything into a formatted Excel workbook

**Data Sources:**
- [ProPublica Nonprofit Explorer API v2](https://projects.propublica.org/nonprofits/api/) (free, no key needed)
- IRS Form 990 XML E-Files (publicly available for electronically-filed returns)
- [Apollo.io](https://apollo.io) for contact enrichment (optional, free tier available)

**Limitations:**
- The search API doesn't support filtering by revenue directly, so we check each org individually
- Schedule G data is only available for e-filed returns (most large orgs e-file)
- API rate limiting: ~0.5-1.5s between requests to be respectful of the free service
    """)
